"""
excel_exporter.py
──────────────────────────────────────────────────────────────────
Signal History tablosunu okuyan ve güzel formatlı bir Excel dosyası
oluşturan modül.

Kullanım (tek seferlik):
    python src/excel_exporter.py

Ana worker'dan otomatik çağrı:
    from excel_exporter import export_signal_history
    export_signal_history()
──────────────────────────────────────────────────────────────────
"""

import os
import sys
from datetime import datetime

sys.path.append(os.path.dirname(os.path.abspath(__file__)))

try:
    import psycopg2
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"❌ Eksik paket: {e}. Çalıştır: pip install psycopg2-binary openpyxl")
    sys.exit(1)

import config

# ─── Çıktı Dosyası ────────────────────────────────────────────────────────────
OUTPUT_DIR  = os.path.join(os.path.dirname(__file__), '..', 'data')
OUTPUT_FILE = os.path.join(OUTPUT_DIR, 'sinyal_gecmisi.xlsx')

# ─── Renkler ──────────────────────────────────────────────────────────────────
C_HEADER      = "1E3A5F"   # Koyu lacivert — header satırı
C_STRONG_BUY  = "0D4723"   # Koyu yeşil  — STRONG_BUY satırı
C_BUY         = "1A3C20"   # Açık yeşil  — BUY satırı
C_WATCH       = "3D3000"   # Koyu sarı   — WATCH satırı
C_ZEBRA       = "111827"   # Alternatif satır arka planı
C_DARK        = "0B0F19"   # Ana arka plan

C_DIAMOND     = "00B4D8"   # Camgöbeği  — Diamond conviction
C_GOLD        = "FFB700"   # Altın      — Gold conviction
C_SILVER      = "94A3B8"   # Gümüş      — Silver conviction
C_BRONZE      = "CD7F32"   # Bronz      — Bronze conviction

C_BULL        = "22C55E"   # Yeşil — BULL
C_BEAR        = "EF4444"   # Kırmızı — BEAR
C_SIDEWAYS    = "94A3B8"   # Gri — SIDEWAYS

C_PERF_POS    = "166534"   # Koyu yeşil — pozitif perf
C_PERF_NEG    = "7F1D1D"   # Koyu kırmızı — negatif perf
C_TEXT_BRIGHT = "F1F5F9"   # Açık beyaz metin
C_TEXT_DIM    = "64748B"   # Soluk metin

# ─── Yardımcı Fonksiyonlar ────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=C_TEXT_BRIGHT, size=9):
    return Font(bold=bold, color=color, name="Calibri", size=size)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=False)

def _right():
    return Alignment(horizontal="right", vertical="center")

def _left():
    return Alignment(horizontal="left", vertical="center")

def _border():
    thin = Side(style="thin", color="1E293B")
    return Border(bottom=thin)

def _perf_str(val):
    if val is None:
        return "—"
    p = float(val)
    arrow = "▲" if p >= 0 else "▼"
    return f"{arrow} %{abs(p):.2f}"

def _perf_fill(val):
    if val is None:
        return _fill(C_DARK)
    return _fill(C_PERF_POS) if float(val) >= 0 else _fill(C_PERF_NEG)

# ─── Ana Fonksiyon ────────────────────────────────────────────────────────────
def export_signal_history():
    """PostgreSQL signal_history tablosunu okur, Excel'e yazar."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # DB Bağlantısı
    try:
        conn = psycopg2.connect(**config.DB_AYARLARI)
        cur  = conn.cursor()
    except Exception as e:
        print(f"❌ DB bağlantı hatası: {e}")
        return

    cur.execute("""
        WITH deduped AS (
            SELECT *,
                   ROW_NUMBER() OVER (
                       PARTITION BY sembol, signal, DATE(signal_date)
                       ORDER BY signal_date DESC
                   ) AS rn
            FROM signal_history
        )
        SELECT
            sembol, signal_date, signal, conviction, score, unified_score,
            fiyat, stop_price, target_price,
            CASE WHEN target_price > 0 AND stop_price > 0 AND fiyat > 0
                 THEN ROUND((target_price - fiyat) / NULLIF(fiyat - stop_price, 0), 2)
                 ELSE NULL END AS rr_ratio,
            rsi, adx, macd_hist,
            market_regime, main_strategy,
            ARRAY_TO_STRING(tags, ', ') AS tags_str,
            fiyat_1gun, perf_1gun,
            fiyat_1hafta, perf_1hafta,
            fiyat_1ay, perf_1ay
        FROM deduped
        WHERE rn = 1
        ORDER BY signal_date DESC
        LIMIT 1000
    """)
    rows = cur.fetchall()
    conn.close()

    if not rows:
        print("⚠️ signal_history tablosu boş. Henüz kayıt yok.")
        return

    # ─── Workbook Oluştur ─────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sinyal Geçmişi"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B3"  # İlk 2 satır + sembol sütunu sabit

    # ─── Başlık Satırı (Grup Başlıkları) ─────────────────────────────────────
    grup_baslik = [
        (1,  1,  "📋 TANIM"),
        (7,  9,  "💰 FİYAT BİLGİSİ"),
        (10, 10, "R/R"),
        (11, 13, "📊 TEKNİK"),
        (14, 15, "🌍 PİYASA"),
        (16, 16, "🏷️ TAGS"),
        (17, 18, "📅 +1 GÜN"),
        (19, 20, "📅 +1 HAFTA"),
        (21, 22, "📅 +1 AY"),
    ]

    ws.row_dimensions[1].height = 22
    for start_col, end_col, label in grup_baslik:
        cell = ws.cell(row=1, column=start_col, value=label)
        cell.fill      = _fill("0F172A")
        cell.font      = _font(bold=True, size=8, color="94A3B8")
        cell.alignment = _center()
        if end_col > start_col:
            ws.merge_cells(
                start_row=1, start_column=start_col,
                end_row=1, end_column=end_col
            )

    # ─── Sütun Başlıkları ─────────────────────────────────────────────────────
    headers = [
        # Tanım
        "Sembol", "Tarih", "Sinyal", "Conviction", "Score", "Unified",
        # Fiyat
        "Giriş Fiyatı", "Stop", "Hedef",
        # R/R
        "Risk/Ödül",
        # Teknik
        "RSI", "ADX", "MACD Hist",
        # Piyasa
        "Regime", "Strateji",
        # Tags
        "Tags",
        # +1 Gün
        "T+1 Fiyat", "T+1 %",
        # +1 Hafta
        "T+5 Fiyat", "T+5 %",
        # +1 Ay
        "T+21 Fiyat", "T+21 %",
    ]

    ws.row_dimensions[2].height = 28
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.fill      = _fill(C_HEADER)
        cell.font      = _font(bold=True, size=10)
        cell.alignment = _center()
        cell.border    = _border()

    # ─── Sütun Genişlikleri ───────────────────────────────────────────────────
    col_widths = [
        10, 18, 14, 12, 8, 9,   # Tanım
        13, 12, 12,              # Fiyat
        9,                       # R/R
        8, 8, 10,                # Teknik
        10, 12,                  # Piyasa
        35,                      # Tags
        13, 10,                  # +1 Gün
        13, 10,                  # +1 Hafta
        13, 10,                  # +1 Ay
    ]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # ─── Veri Satırları ───────────────────────────────────────────────────────
    for row_idx, row in enumerate(rows, 3):
        (sembol, signal_date, signal, conviction, score, unified_score,
         fiyat, stop_price, target_price, rr_ratio,
         rsi, adx, macd_hist,
         market_regime, main_strategy, tags_str,
         fiyat_1gun, perf_1gun,
         fiyat_1hafta, perf_1hafta,
         fiyat_1ay, perf_1ay) = row

        # Satır arka plan rengi (sinyal tipine göre)
        if signal == "STRONG_BUY":
            row_fill = _fill(C_STRONG_BUY)
        elif signal == "BUY":
            row_fill = _fill(C_BUY)
        elif signal == "WATCH":
            row_fill = _fill(C_WATCH)
        else:
            row_fill = _fill(C_ZEBRA if row_idx % 2 == 0 else C_DARK)

        ws.row_dimensions[row_idx].height = 20

        def put(col, value, align=None, fill=None, font=None, num_fmt=None):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.fill      = fill or row_fill
            cell.font      = font or _font()
            cell.alignment = align or _center()
            cell.border    = _border()
            if num_fmt:
                cell.number_format = num_fmt
            return cell

        # Sembol
        put(1, sembol, _left(), font=_font(bold=True, size=10))

        # Tarih
        tarih_str = signal_date.strftime("%d/%m/%Y %H:%M") if signal_date else "—"
        put(2, tarih_str, _center(), font=_font(color=C_TEXT_DIM, size=8))

        # Sinyal
        sig_labels = {"STRONG_BUY": "🚀 GÜÇLÜ AL", "BUY": "✅ AL", "WATCH": "👀 İZLE"}
        put(3, sig_labels.get(signal, signal), font=_font(bold=True, size=9))

        # Conviction (renkli)
        conv_colors = {"DIAMOND": C_DIAMOND, "GOLD": C_GOLD, "SILVER": C_SILVER, "BRONZE": C_BRONZE}
        conv_labels = {"DIAMOND": "💎 DIAMOND", "GOLD": "🥇 GOLD", "SILVER": "🥈 SILVER", "BRONZE": "🥉 BRONZE"}
        put(4, conv_labels.get(conviction, conviction),
            font=_font(bold=True, color=conv_colors.get(conviction, C_TEXT_BRIGHT)))

        # Score & Unified Score
        score_color = "22C55E" if (score or 0) >= 70 else ("FCD34D" if (score or 0) >= 50 else "EF4444")
        put(5, score or 0, font=_font(bold=True, color=score_color, size=11))
        uni_color = "22C55E" if (unified_score or 0) >= 70 else ("FCD34D" if (unified_score or 0) >= 50 else "EF4444")
        put(6, unified_score or 0, font=_font(bold=True, color=uni_color, size=11))

        # Fiyat, Stop, Hedef
        put(7, float(fiyat or 0), _right(), num_fmt='#,##0.00" ₺"')
        put(8, float(stop_price or 0), _right(), font=_font(color="FCA5A5"), num_fmt='#,##0.00" ₺"')
        put(9, float(target_price or 0), _right(), font=_font(color="86EFAC"), num_fmt='#,##0.00" ₺"')

        # Risk/Ödül
        rr_val = float(rr_ratio) if rr_ratio is not None else None
        rr_str = f"{rr_val:.2f}" if rr_val is not None else "—"
        rr_color = "22C55E" if (rr_val or 0) >= 2 else C_TEXT_DIM
        put(10, rr_str, font=_font(bold=True, color=rr_color))

        # RSI
        rsi_f = float(rsi or 0)
        rsi_color = "22C55E" if rsi_f < 30 else ("EF4444" if rsi_f > 70 else C_TEXT_BRIGHT)
        put(11, round(rsi_f, 1), font=_font(color=rsi_color))

        # ADX
        adx_f = float(adx or 0)
        adx_color = "F9FAFB" if adx_f > 25 else C_TEXT_DIM
        put(12, round(adx_f, 1), font=_font(bold=adx_f > 25, color=adx_color))

        # MACD Hist
        macd_f = float(macd_hist or 0)
        put(13, round(macd_f, 4), font=_font(color="22C55E" if macd_f > 0 else "EF4444"))

        # Regime (renkli)
        reg_colors = {"BULL": C_BULL, "BEAR": C_BEAR, "SIDEWAYS": C_SIDEWAYS}
        reg_labels = {"BULL": "🐂 BULL", "BEAR": "🐻 BEAR", "SIDEWAYS": "↔️ SIDEWAYS"}
        put(14, reg_labels.get(market_regime, market_regime),
            font=_font(bold=True, color=reg_colors.get(market_regime, C_TEXT_DIM)))

        # Strateji
        put(15, main_strategy or "—", font=_font(color=C_TEXT_DIM))

        # Tags
        put(16, tags_str or "—", _left(), font=_font(color=C_TEXT_DIM, size=8))

        # +1 Gün
        put(17, float(fiyat_1gun) if fiyat_1gun else "—", _right())
        perf_cell = ws.cell(row=row_idx, column=18)
        perf_cell.value     = _perf_str(perf_1gun)
        perf_cell.fill      = _perf_fill(perf_1gun)
        perf_cell.font      = _font(bold=True if perf_1gun is not None else False)
        perf_cell.alignment = _center()
        perf_cell.border    = _border()

        # +1 Hafta
        put(19, float(fiyat_1hafta) if fiyat_1hafta else "—", _right())
        perf_cell2 = ws.cell(row=row_idx, column=20)
        perf_cell2.value     = _perf_str(perf_1hafta)
        perf_cell2.fill      = _perf_fill(perf_1hafta)
        perf_cell2.font      = _font(bold=True if perf_1hafta is not None else False)
        perf_cell2.alignment = _center()
        perf_cell2.border    = _border()

        # +1 Ay
        put(21, float(fiyat_1ay) if fiyat_1ay else "—", _right())
        perf_cell3 = ws.cell(row=row_idx, column=22)
        perf_cell3.value     = _perf_str(perf_1ay)
        perf_cell3.fill      = _perf_fill(perf_1ay)
        perf_cell3.font      = _font(bold=True if perf_1ay is not None else False)
        perf_cell3.alignment = _center()
        perf_cell3.border    = _border()

    # ─── Özet Sayfası ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("📊 Özet")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["C"].width = 15
    ws2.column_dimensions["D"].width = 15

    # Özet başlık
    ozet_header = ws2.cell(row=1, column=1, value="📊 Sinyal Performans Özeti")
    ozet_header.font = Font(bold=True, size=14, color=C_TEXT_BRIGHT, name="Calibri")
    ozet_header.fill = _fill(C_HEADER)
    ws2.merge_cells("A1:D1")
    ws2.row_dimensions[1].height = 30
    ozet_header.alignment = _center()

    # Zaman bilgisi
    ws2.cell(row=2, column=1, value=f"Son güncelleme: {datetime.now().strftime('%d.%m.%Y %H:%M')}").font = _font(color=C_TEXT_DIM)

    # Tablonun özeti
    rows_arr = list(rows)
    total     = len(rows_arr)
    strong    = sum(1 for r in rows_arr if r[2] == "STRONG_BUY")
    buy_c     = sum(1 for r in rows_arr if r[2] == "BUY")
    watch_c   = sum(1 for r in rows_arr if r[2] == "WATCH")

    def perf_avg(col_idx):
        vals = [float(r[col_idx]) for r in rows_arr if r[col_idx] is not None]
        return (sum(vals) / len(vals)) if vals else None

    avg1g = perf_avg(17)   # perf_1gun
    avg1h = perf_avg(19)   # perf_1hafta
    avg1a = perf_avg(21)   # perf_1ay

    summary_rows = [
        (4,  "Metrik",          "Değer"),
        (5,  "Toplam Sinyal",   total),
        (6,  "🚀 Strong Buy",   strong),
        (7,  "✅ Buy",          buy_c),
        (8,  "👀 Watch",        watch_c),
        (9,  "Ort. +1 Gün %",  f"{'▲' if (avg1g or 0) >= 0 else '▼'} %{abs(avg1g or 0):.2f}" if avg1g is not None else "—"),
        (10, "Ort. +1 Hafta %", f"{'▲' if (avg1h or 0) >= 0 else '▼'} %{abs(avg1h or 0):.2f}" if avg1h is not None else "—"),
        (11, "Ort. +1 Ay %",   f"{'▲' if (avg1a or 0) >= 0 else '▼'} %{abs(avg1a or 0):.2f}" if avg1a is not None else "—"),
    ]

    for r_idx, label, val in summary_rows:
        ws2.row_dimensions[r_idx].height = 22
        is_header = (r_idx == 4)
        for c_idx, cell_val in enumerate([label, val], 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=cell_val)
            cell.fill = _fill(C_HEADER if is_header else (C_DARK if r_idx % 2 == 0 else C_ZEBRA))
            cell.font = _font(bold=is_header, size=10 if is_header else 9)
            cell.alignment = _right() if c_idx == 2 else _left()
            cell.border = _border()

    # ─── Kaydet ───────────────────────────────────────────────────────────────
    wb.save(OUTPUT_FILE)
    print(f"✅ Excel kaydedildi: {os.path.abspath(OUTPUT_FILE)} ({total} sinyal)")
    return OUTPUT_FILE


if __name__ == "__main__":
    print("📊 Sinyal geçmişi Excel'e aktarılıyor...")
    export_signal_history()
